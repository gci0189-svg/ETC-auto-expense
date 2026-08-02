"""
DN 費用申報整合工具 v4
"""

import streamlit as st
import streamlit.components.v1 as components
import openpyxl
import pdfplumber
try:
    import fitz
    FITZ_AVAILABLE = True
except ImportError:
    FITZ_AVAILABLE = False
import io, os, re, math
import json
import pandas as pd
import shutil
import tempfile
import subprocess
from datetime import datetime
from PIL import Image

try:
    from pyzbar.pyzbar import decode as decode_qrcode
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False

try:
    from pypdf import PdfReader, PdfWriter
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False

SOFFICE_PATH = shutil.which('soffice') or shutil.which('libreoffice')
LIBREOFFICE_AVAILABLE = SOFFICE_PATH is not None

st.set_page_config(page_title="DN 費用申報整合工具", layout="wide", page_icon="🚗")

st.markdown("""
<style>
  .block-container{padding-top:1rem;padding-bottom:1rem;padding-left:1rem;padding-right:1rem;max-width:1280px!important;margin:0 auto!important}
  section[data-testid="stMain"] > div {padding-left:1rem}
  h1,h2,h3{margin-top:0}
  h2{font-size:1.15rem!important;color:#1F4E79;border-bottom:2px solid #1F4E79;padding-bottom:4px}
  h3{font-size:1rem!important;color:#333}
  .success-box{background:#E8F5E9;border-left:4px solid #2E7D32;padding:.6rem 1rem;border-radius:4px;margin:.4rem 0;font-size:.9rem}
  .warn-box{background:#FFF8E1;border-left:4px solid #F59E0B;padding:.6rem 1rem;border-radius:4px;margin:.4rem 0;font-size:.9rem}
  .info-box{background:#E8F4FD;border-left:4px solid #1F4E79;padding:.6rem 1rem;border-radius:4px;margin:.4rem 0;font-size:.9rem}
  .section-title{font-size:1.05rem;font-weight:700;color:#1F4E79;border-bottom:2px solid #1F4E79;padding-bottom:4px;margin-bottom:.8rem}
  [data-testid="stHorizontalBlock"] {flex-direction:row!important;flex-wrap:nowrap!important;gap:1.2rem!important;}
  [data-testid="column"] {min-width:0!important;}
  @media (max-width:640px){div[data-testid="stHorizontalBlock"]{flex-direction:row!important;flex-wrap:nowrap!important;}}
</style>
""", unsafe_allow_html=True)

STATE_DIR = ".state"

def save_persistent_state():
    try:
        os.makedirs(STATE_DIR, exist_ok=True)
        config = {}
        for k in ['mileage_allowance','selected_sheet','mileage_manual',
                  'tolls_parking_amount','mileage_distance','fuel_amount','fuel_tax']:
            config[k] = st.session_state.get(k)
        for i in range(1, 11):
            config[f"inv_t{i}"] = st.session_state.get(f"inv_t{i}", 0)
            config[f"inv_x{i}"] = st.session_state.get(f"inv_x{i}", 0)
        with open(os.path.join(STATE_DIR, "config.json"), "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        bin_files = {
            "toll_excel.xlsx": st.session_state.get("toll_excel"),
            "toll_pdf_out.pdf": st.session_state.get("toll_pdf_out"),
            "mileage_pdf_out.pdf": st.session_state.get("mileage_pdf_out"),
            "merged_pdf.pdf": st.session_state.get("merged_pdf"),
            "telecom_pdf.pdf": st.session_state.get("telecom_pdf")
        }
        for fname, content in bin_files.items():
            fpath = os.path.join(STATE_DIR, fname)
            if content is not None:
                with open(fpath, "wb") as f:
                    f.write(content)
            elif os.path.exists(fpath):
                os.remove(fpath)
        audit_df = st.session_state.get("audit_df")
        audit_path = os.path.join(STATE_DIR, "audit_df.csv")
        if audit_df is not None:
            audit_df.to_csv(audit_path, index=False)
        elif os.path.exists(audit_path):
            os.remove(audit_path)
    except Exception:
        pass

def load_persistent_state():
    if st.session_state.get("state_loaded"):
        return
    try:
        config_path = os.path.join(STATE_DIR, "config.json")
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)
            for k, v in config.items():
                st.session_state[k] = v
            bin_files = {
                "toll_excel.xlsx": "toll_excel",
                "toll_pdf_out.pdf": "toll_pdf_out",
                "mileage_pdf_out.pdf": "mileage_pdf_out",
                "merged_pdf.pdf": "merged_pdf",
                "telecom_pdf.pdf": "telecom_pdf"
            }
            for fname, state_key in bin_files.items():
                fpath = os.path.join(STATE_DIR, fname)
                if os.path.exists(fpath):
                    with open(fpath, "rb") as f:
                        st.session_state[state_key] = f.read()
            audit_path = os.path.join(STATE_DIR, "audit_df.csv")
            if os.path.exists(audit_path):
                st.session_state.audit_df = pd.read_csv(audit_path)
        st.session_state.state_loaded = True
    except Exception:
        pass

def clear_persistent_state():
    try:
        if os.path.exists(STATE_DIR):
            shutil.rmtree(STATE_DIR)
    except Exception:
        pass
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

for k in ['toll_excel','toll_pdf_out','telecom_pdf','mileage_allowance',
          'selected_sheet','mileage_manual','merged_pdf','audit_df','mileage_pdf_out',
          'tolls_parking_amount','mileage_distance','fuel_amount','fuel_tax']:
    if k not in st.session_state:
        st.session_state[k] = None if k not in ['mileage_manual','tolls_parking_amount','mileage_distance','fuel_amount','fuel_tax'] else 0

load_persistent_state()

for i in range(1, 11):
    if f"inv_t{i}" not in st.session_state:
        st.session_state[f"inv_t{i}"] = 0
    if f"inv_x{i}" not in st.session_state:
        st.session_state[f"inv_x{i}"] = 0

invoice_rows = []
mileage_input = 0

t_col1, t_col2 = st.columns([4, 1])
with t_col1:
    st.markdown('<p style="font-size:1rem;font-weight:700;color:#1F4E79;margin:0 0 0.8rem 0;">🚗 DN 費用申報整合工具</p>', unsafe_allow_html=True)
with t_col2:
    if st.button("🔄 一鍵重置 / 新月份對帳", type="secondary", use_container_width=True):
        clear_persistent_state()


def auto_tax(i):
    total = st.session_state[f"inv_t{i}"]
    if total > 0:
        sales = round(total / 1.05)
        st.session_state[f"inv_x{i}"] = round(sales * 0.05)
    else:
        st.session_state[f"inv_x{i}"] = 0


def format_date_slash(v):
    try:
        if isinstance(v, str):
            return pd.to_datetime(v.strip()).strftime('%Y/%m/%d')
        if hasattr(v, 'strftime'):
            return v.strftime('%Y/%m/%d')
    except Exception:
        pass
    return None


def read_mileage_allowance(excel_bytes, sheet_name):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if any(str(v).strip() == '小計' for v in vals if v is not None):
            for idx in [9, 10, 8, 11, 7]:
                if idx < len(vals) and vals[idx] is not None:
                    try:
                        v = float(vals[idx])
                        if v > 0:
                            return v
                    except (TypeError, ValueError):
                        pass
    return None


def parse_fuel_pdf_totals(pdf_bytes):
    """
    [快速精確版 - CT欄位定位法]：
    1. 用 CT- 發票號碼定位各欄 X 邊界（0.15s，比空間投影版快 5-10x）
    2. 找 TX 關鍵字左側金額，對應欄位
    3. 去重用 (日期, 金額) 組合，允許同金額不同日期
    """
    results = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words()

            # 1. 找發票欄位（CT- 開頭），按 X 軸排序
            ct_words = sorted(
                [w for w in words if re.match(r'^CT-\d+$', w['text'])],
                key=lambda w: w['x0']
            )
            if not ct_words:
                continue

            # 2. 建欄位 X 中點邊界
            midpoints = []
            for i in range(len(ct_words) - 1):
                midpoints.append((ct_words[i]['x1'] + ct_words[i + 1]['x0']) / 2)

            def get_col(x):
                for i, m in enumerate(midpoints):
                    if x < m:
                        return i
                return len(ct_words) - 1

            # 3. 找 TX → 左側最近金額 → 欄位索引
            col_amounts = {}
            tx_words = [w for w in words if w['text'].upper() == 'TX']
            for tx in tx_words:
                same_row = [w for w in words
                            if abs(w['top'] - tx['top']) < 8 and w['x0'] < tx['x0']]
                candidates = [w for w in same_row
                              if re.match(r'^\d{3,5}$', w['text']) and 500 <= int(w['text']) <= 9999]
                if not candidates:
                    continue
                amt_w = max(candidates, key=lambda w: w['x0'])
                col = get_col((amt_w['x0'] + amt_w['x1']) / 2)
                col_amounts[col] = int(amt_w['text'])

            # 4. 找各欄日期
            date_words = [w for w in words if re.match(r'^20\d{2}-\d{2}-\d{2}$', w['text'])]
            col_dates = {}
            for dw in date_words:
                col = get_col((dw['x0'] + dw['x1']) / 2)
                col_dates[col] = dw['text'].replace('-', '/')

            # 5. 日期不足時從 extract_text 補
            if len(col_dates) < len(col_amounts):
                text = page.extract_text() or ""
                found = re.findall(r'(20\d{2}-\d{2}-\d{2})', text)
                for col in col_amounts:
                    if col not in col_dates:
                        col_dates[col] = found[0].replace('-', '/') if found else "9999/12/31"

            # 6. 組合 (日期, 金額)
            for col, amt in col_amounts.items():
                results.append((col_dates.get(col, "9999/12/31"), amt))

    # 依日期排序，(日期, 金額) 組合去重（允許同金額不同日期）
    seen = set()
    final = []
    for d, a in sorted(results):
        key = (d, a)
        if key not in seen:
            seen.add(key)
            final.append(a)
    return final


def parse_toll_from_pdf(pdf_bytes):
    toll_map = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            rows = re.findall(r'(\d{4}/\d{2}/\d{2})\s+([\d\.]+)(?:公里)?\s+(\d+)(?:元)?', text)
            for date_str, mileage, amt in rows:
                std_date = format_date_slash(date_str)
                if std_date:
                    toll_map[std_date] = toll_map.get(std_date, 0) + int(amt)
    return toll_map


def find_font():
    for f in os.listdir("."):
        if f.lower().endswith((".ttc", ".ttf")):
            return f
    for root, dirs, files in os.walk("."):
        for f in files:
            if f.lower().endswith((".ttc", ".ttf")):
                return os.path.join(root, f)
    for fp in ['/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
               '/System/Library/Fonts/PingFang.ttc',
               'C:/Windows/Fonts/msjh.ttc']:
        if os.path.exists(fp):
            return fp
    return None


def install_local_fonts():
    try:
        user_font_dir = os.path.expanduser('~/.fonts')
        if not os.path.exists(user_font_dir):
            os.makedirs(user_font_dir)
        fonts_copied = False
        for f in os.listdir('.'):
            if f.lower().endswith(('.ttf', '.ttc')):
                dest_path = os.path.join(user_font_dir, f)
                if not os.path.exists(dest_path):
                    shutil.copy(f, dest_path)
                    fonts_copied = True
        if fonts_copied:
            subprocess.run(['fc-cache', '-f'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
    except Exception:
        pass


def convert_excel_to_pdf(excel_bytes, sheet_name):
    if not LIBREOFFICE_AVAILABLE:
        return None
    try:
        install_local_fonts()
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]
        ws = wb[sheet_name]
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = '9'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        wb.active = 0
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_xlsx:
            xlsx_path = tmp_xlsx.name
            wb.save(xlsx_path)
        output_dir = tempfile.gettempdir()
        cmd = [SOFFICE_PATH, '--headless', '--convert-to', 'pdf', '--outdir', output_dir, xlsx_path]
        subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
        pdf_filename = os.path.basename(xlsx_path).replace('.xlsx', '.pdf')
        pdf_path = os.path.join(output_dir, pdf_filename)
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        try:
            os.remove(xlsx_path)
            os.remove(pdf_path)
        except:
            pass
        return pdf_bytes
    except Exception as e:
        st.error(f"PDF 轉換失敗: {e}")
        return None


def remove_pdf_password_and_extract_page1(pdf_bytes, password=""):
    if FITZ_AVAILABLE:
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            total_pages = len(doc)
            if doc.is_encrypted:
                if not doc.authenticate(password):
                    doc.close()
                    return False, None, f"密碼錯誤（嘗試：「{password}」）"
            new_doc = fitz.open()
            new_doc.insert_pdf(doc, from_page=0, to_page=0)
            out = io.BytesIO()
            new_doc.save(out, encryption=fitz.PDF_ENCRYPT_NONE)
            new_doc.close(); doc.close()
            return True, out.getvalue(), f"成功！已移除密碼並擷取第1頁（共 {total_pages} 頁）"
        except Exception:
            pass
    if PYPDF_AVAILABLE:
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            if reader.is_encrypted:
                if reader.decrypt(password) == 0:
                    return False, None, f"密碼錯誤（嘗試：「{password}」）"
            writer = PdfWriter()
            writer.add_page(reader.pages[0])
            out = io.BytesIO()
            writer.write(out)
            return True, out.getvalue(), f"成功（備援）！擷取第1頁（共 {len(reader.pages)} 頁）"
        except Exception as e:
            return False, None, f"處理失敗：{e}"
    return False, None, "解密失敗，請確認密碼"


def safe_format_num(val):
    if val is None:
        return "0"
    try:
        return f"{int(float(val)):,}"
    except Exception:
        return "0"


def build_results_html(invoice_rows, mileage_allowance):
    total_amount = sum(r[0] for r in invoice_rows)
    total_tax    = sum(r[1] for r in invoice_rows)
    km = math.ceil(max(0, mileage_allowance - total_amount) / 7) if mileage_allowance > 0 else 0
    amt = km * 7

    TD    = "border:1px solid #bbb;padding:4px 6px;font-size:11px;font-family:Arial,sans-serif;"
    TDNUM = TD + "text-align:right;"
    HDR   = TD + "background:#1F4E79;color:#fff;font-weight:700;text-align:center;font-size:12px;"
    SUB   = TD + "background:#D6E4F0;font-weight:700;text-align:center;"
    TOT   = TD + "background:#BDD7EE;font-weight:700;"
    BLK   = "border:none;background:transparent;width:12px;"

    right_rows = [
        ("總里程津貼",          f"{int(mileage_allowance):,}" if mileage_allowance else "—", "#FFF2CC", "#1F4E79", True),
        ("加油發票合計",         f"{total_amount:,}",  "#FFFFFF", "#333", False),
        ("發票稅額合計",         f"{total_tax:,}",     "#FCE4D6", "#C00000", False),
        ("Personal Car 公里數",  f"{km:,}",            "#E2EFDA", "#C00000", True),
        ("Personal Car 金額",    f"{amt:,}",           "#E2EFDA", "#333", False),
        ("Fuel（油資補助）",     f"{total_amount:,}",  "#FFFFFF", "#333", False),
    ]

    rows_html = ""
    for i in range(10):
        l1 = datetime.now().strftime('%Y/%m') if i < len(invoice_rows) else ""
        l2 = f"{invoice_rows[i][0]:,}"        if i < len(invoice_rows) else ""
        l3 = f"{invoice_rows[i][1]:,}"        if i < len(invoice_rows) else ""
        if i < len(right_rows):
            rl, rv, rbg, rc, rb = right_rows[i]
            fw = "700" if rb else "400"
            fs = "12px" if rb else "11px"
            r1 = f'<td style="{TD}background:{rbg};">{rl}</td>'
            r2 = f'<td style="{TDNUM}background:{rbg};color:{rc};font-weight:{fw};font-size:{fs};">{rv}</td>'
        else:
            r1 = f'<td style="{TD}"></td>'
            r2 = f'<td style="{TD}"></td>'
        rows_html += (
            f'<tr>'
            f'<td style="{TD}">{l1}</td>'
            f'<td style="{TDNUM}">{l2}</td>'
            f'<td style="{TDNUM}">{l3}</td>'
            f'<td style="{BLK}"></td>'
            f'{r1}{r2}</tr>'
        )

    formula = (
        f"⌈({int(mileage_allowance):,} − {total_amount:,}) ÷ 7⌉ = {km:,} 公里"
        if mileage_allowance else ""
    )

    html = (
        '<div style="overflow-x:auto;">'
        '<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;">'
        '<colgroup>'
        '<col style="width:13%"><col style="width:14%"><col style="width:12%">'
        '<col style="width:2%">'
        '<col style="width:35%"><col style="width:18%">'
        '</colgroup>'
        f'<tr><td colspan="3" style="{HDR}">加油發票登記</td>'
        f'<td style="{BLK}"></td>'
        f'<td colspan="2" style="{HDR}">申報金額計算</td></tr>'
        f'<tr><td style="{SUB}">日期</td><td style="{SUB}">發票總額</td>'
        f'<td style="{SUB}">發票稅額</td><td style="{BLK}"></td>'
        f'<td style="{SUB}">項目</td><td style="{SUB}">金額 (TWD)</td></tr>'
        + rows_html +
        f'<tr><td style="{TOT}">合計</td>'
        f'<td style="{TOT}text-align:right;">{total_amount:,}</td>'
        f'<td style="{TOT}text-align:right;">{total_tax:,}</td>'
        f'<td style="{BLK}"></td>'
        f'<td colspan="2" style="{TD}font-size:10px;color:#666;">{formula}</td></tr>'
        '</table></div>'
    )
    return html, total_amount, total_tax, km, amt


# ═══════════════════════════════════════════
# 主要佈局
# ═══════════════════════════════════════════
col_toll, col_fuel = st.columns([1.2, 1], gap="large")

# ╔══════════════════════════════════════════╗
# ║  左側：通行費對帳                        ║
# ╚══════════════════════════════════════════╝
with col_toll:
    st.markdown('<div class="section-title">🛣️ 通行費對帳</div>', unsafe_allow_html=True)

    parking_pdf = st.file_uploader("① 停車費 PDF", type="pdf", key="parking_pdf")
    toll_pdf    = st.file_uploader("② 遠通電收 PDF", type="pdf", key="toll_pdf")
    te_excel    = st.file_uploader("③ T_E 申請表 (.xlsx)", type="xlsx", key="te_main")

    selected_sheet = None
    if te_excel:
        wb_tmp = openpyxl.load_workbook(te_excel, read_only=True)
        sheets = wb_tmp.sheetnames
        cm = f"{datetime.now().month}月"
        default_idx = sheets.index(cm) if cm in sheets else 0
        selected_sheet = st.selectbox("④ 選擇月份工作表", sheets, index=default_idx, key="s_main")
        st.session_state.selected_sheet = selected_sheet

        if selected_sheet:
            te_excel.seek(0)
            excel_bytes = te_excel.read()
            allowance = read_mileage_allowance(excel_bytes, selected_sheet)
            if allowance:
                st.session_state.mileage_allowance = allowance
                st.markdown(f"""
                <div class="success-box">
                ✅ <b>{selected_sheet}</b> 里程津貼小計：<b>NT$ {int(allowance):,}</b>
                （已同步至右側加油費計算）
                </div>""", unsafe_allow_html=True)

            active_bytes = st.session_state.toll_excel if st.session_state.toll_excel is not None else excel_bytes
            try:
                wb_auto = openpyxl.load_workbook(io.BytesIO(active_bytes), data_only=True)
                if selected_sheet in wb_auto.sheetnames:
                    ws_auto = wb_auto[selected_sheet]
                    t_total = 0
                    p_total = 0
                    for r in range(8, ws_auto.max_row + 1):
                        is_sub = False
                        for col_idx in [1, 2, 3]:
                            val = ws_auto.cell(row=r, column=col_idx).value
                            if val and str(val).strip() == "小計":
                                is_sub = True
                                break
                        if is_sub:
                            continue
                        val_k = ws_auto.cell(row=r, column=11).value
                        val_l = ws_auto.cell(row=r, column=12).value
                        try:
                            if val_k is not None: t_total += int(float(val_k))
                        except: pass
                        try:
                            if val_l is not None: p_total += int(float(val_l))
                        except: pass
                    st.session_state.tolls_parking_amount = t_total + p_total
            except:
                pass

    if toll_pdf and te_excel and selected_sheet:
        if st.button("🚀 開始對帳與標註", type="primary", key="run_toll"):
            with st.spinner("對帳比對、標註中..."):
                try:
                    toll_pdf.seek(0)
                    toll_map = parse_toll_from_pdf(toll_pdf.read())
                    if not toll_map:
                        st.error("無法解析通行費PDF，請確認格式")
                        st.stop()

                    te_excel.seek(0)
                    wb = openpyxl.load_workbook(te_excel)
                    ws = wb[selected_sheet]
                    DATE_COL, TOLL_COL, ITEM_COL = 4, 11, 1
                    serial_map, matched = {}, set()

                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                    ws.page_setup.paperSize = '9'
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0

                    for row in range(8, ws.max_row + 1):
                        raw_date = ws.cell(row=row, column=DATE_COL).value
                        if not raw_date: continue
                        d_str = format_date_slash(raw_date)
                        if not d_str: continue
                        if d_str in toll_map and d_str not in matched:
                            ws.cell(row=row, column=TOLL_COL).value = toll_map[d_str]
                            item_val = ws.cell(row=row, column=ITEM_COL).value
                            if item_val is not None:
                                try:    serial_map[d_str] = f"項目 {int(float(item_val)):02d}"
                                except: serial_map[d_str] = f"項目 {item_val}"
                            matched.add(d_str)

                    excel_daily = {}
                    for row in range(8, ws.max_row + 1):
                        raw_date = ws.cell(row=row, column=DATE_COL).value
                        if not raw_date: continue
                        d_str = format_date_slash(raw_date)
                        if not d_str: continue
                        val = ws.cell(row=row, column=TOLL_COL).value
                        val_num = 0
                        if val is not None:
                            try:    val_num = int(float(val))
                            except: pass
                        excel_daily[d_str] = excel_daily.get(d_str, 0) + val_num

                    all_dates = sorted(list(set(excel_daily.keys()) | set(toll_map.keys())))
                    audit_rows = []
                    for d in all_dates:
                        ex_val = excel_daily.get(d, 0)
                        pdf_val = toll_map.get(d, 0)
                        diff = ex_val - pdf_val
                        status = "✅ 匹配" if diff == 0 else "❌ 金額不符"
                        audit_rows.append({"日期": d, "Excel金額": ex_val, "PDF金額": pdf_val, "差異": diff, "狀態": status})

                    st.session_state.audit_df = pd.DataFrame(audit_rows)

                    tolls_total = 0
                    parking_total = 0
                    for r in range(8, ws.max_row + 1):
                        is_sub_col = False
                        for col_idx in [1, 2, 3]:
                            val = ws.cell(row=r, column=col_idx).value
                            if val and str(val).strip() == "小計":
                                is_sub_col = True
                                break
                        if is_sub_col:
                            continue
                        val_k = ws.cell(row=r, column=11).value
                        val_l = ws.cell(row=r, column=12).value
                        try:
                            if val_k is not None: tolls_total += int(float(val_k))
                        except: pass
                        try:
                            if val_l is not None: parking_total += int(float(val_l))
                        except: pass
                    st.session_state.tolls_parking_amount = tolls_total + parking_total

                    audit_sheet_name = f"對帳稽核_{selected_sheet}"
                    if audit_sheet_name in wb.sheetnames:
                        del wb[audit_sheet_name]
                    audit_ws = wb.create_sheet(title=audit_sheet_name)
                    headers = ["日期", "Excel金額", "PDF金額", "差異", "狀態"]
                    audit_ws.append(headers)
                    for col_num, header in enumerate(headers, 1):
                        cell = audit_ws.cell(row=1, column=col_num)
                        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                        cell.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                    for r in audit_rows:
                        audit_ws.append([r["日期"], r["Excel金額"], r["PDF金額"], r["差異"], r["狀態"]])
                    for col in audit_ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        audit_ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                    out_excel = io.BytesIO()
                    wb.save(out_excel)
                    excel_saved_bytes = out_excel.getvalue()
                    st.session_state.toll_excel = excel_saved_bytes

                    if LIBREOFFICE_AVAILABLE:
                        st.session_state.mileage_pdf_out = convert_excel_to_pdf(excel_saved_bytes, selected_sheet)

                    # ── PDF 標註 ──
                    font_path = find_font()
                    toll_pdf.seek(0)
                    pdf_raw = toll_pdf.read()

                    if FITZ_AVAILABLE:
                        doc = fitz.open(stream=pdf_raw, filetype="pdf")
                        for page in doc:
                            words = page.get_text("words")
                            if font_path:
                                try:   page.insert_font(fontname="cf", fontfile=font_path)
                                except: font_path = None
                            for w in words:
                                if w[4] not in serial_map: continue
                                dw = w
                                lw = sorted([x for x in words if abs(x[1]-dw[1]) < 5], key=lambda x: x[0])
                                km_w  = next((x for x in lw if "公里" in x[4]), None)
                                toll_w = lw[lw.index(km_w)+1] if km_w and lw.index(km_w)+1 < len(lw) else None
                                mx = (km_w[2]+toll_w[0])/2 if (km_w and toll_w) else dw[2]+140
                                if font_path:
                                    page.insert_text((mx-18, dw[3]-2), serial_map[w[4]], fontsize=11, fontname="cf", color=(0,0,0.7))
                                else:
                                    page.insert_text((mx-18, dw[3]-2), serial_map[w[4]], fontsize=11, color=(0,0,0.7))
                        out_toll_only = io.BytesIO()
                        doc.save(out_toll_only)
                        annotated_pdf = out_toll_only.getvalue()
                    else:
                        # pypdf 備援
                        from pypdf import PdfReader as _PR, PdfWriter as _PW
                        from pypdf.generic import (
                            ArrayObject as _AO, DictionaryObject as _DO,
                            NameObject as _NO, DecodedStreamObject as _DSO
                        )
                        def _add_helv(writer, page_obj):
                            fd = _DO({_NO("/Type"):_NO("/Font"),_NO("/Subtype"):_NO("/Type1"),
                                      _NO("/BaseFont"):_NO("/Helvetica"),_NO("/Encoding"):_NO("/WinAnsiEncoding")})
                            ref = writer._add_object(fd)
                            if "/Resources" not in page_obj:
                                page_obj[_NO("/Resources")] = _DO()
                            res = page_obj["/Resources"]
                            if hasattr(res,"get_object"): res = res.get_object()
                            if "/Font" not in res: res[_NO("/Font")] = _DO()
                            fd2 = res["/Font"]
                            if hasattr(fd2,"get_object"): fd2 = fd2.get_object()
                            fd2[_NO("/FHelv")] = ref

                        def _stream(anns):
                            ls = ["q"]
                            for x, y, txt, r, g, b in anns:
                                s = txt.replace("\\","\\\\").replace("(","\\(").replace(")","\\)")
                                ls += ["BT", "/FHelv 10 Tf",
                                       f"{r:.2f} {g:.2f} {b:.2f} rg",
                                       f"{x:.2f} {y:.2f} Td",
                                       f"({s}) Tj", "ET"]
                            ls.append("Q")
                            return "\n".join(ls).encode("latin-1")

                        reader2 = _PR(io.BytesIO(pdf_raw))
                        writer2 = _PW()
                        with pdfplumber.open(io.BytesIO(pdf_raw)) as plumb2:
                            for pi2 in range(len(reader2.pages)):
                                writer2.add_page(reader2.pages[pi2])
                                pg2 = writer2.pages[pi2]
                                pp2 = plumb2.pages[pi2]
                                ph2 = float(pp2.height)
                                wds2 = pp2.extract_words()
                                anns2 = []
                                for wd in wds2:
                                    if not re.match(r"\d{4}/\d{2}/\d{2}", wd["text"]): continue
                                    if wd["text"] not in serial_map: continue
                                    rw = [x for x in wds2 if abs(x["top"]-wd["top"])<5]
                                    km2 = next((x for x in rw if "公里" in x["text"]), None)
                                    if not km2: continue
                                    anns2.append((km2["x1"]+5, ph2-wd["bottom"]+2, serial_map[wd["text"]], 0, 0, 0.7))
                                if anns2:
                                    _add_helv(writer2, pg2)
                                    ns2 = _DSO()
                                    ns2.set_data(_stream(anns2))
                                    nr2 = writer2._add_object(ns2)
                                    ex2 = pg2.get("/Contents")
                                    if ex2 is None: pg2[_NO("/Contents")] = nr2
                                    elif hasattr(ex2,"indirect_reference"): pg2[_NO("/Contents")] = _AO([ex2.indirect_reference, nr2])
                                    else: pg2[_NO("/Contents")] = _AO([ex2, nr2])
                        out2 = io.BytesIO()
                        writer2.write(out2)
                        annotated_pdf = out2.getvalue()

                    st.session_state.toll_pdf_out = annotated_pdf

                    # ── 合併停車費 + 標註遠通 ──
                    SIZE_LIMIT = 15 * 1024 * 1024
                    if parking_pdf and FITZ_AVAILABLE:
                        parking_pdf.seek(0)
                        parking_doc = fitz.open(stream=parking_pdf.read(), filetype="pdf")
                        toll_doc    = fitz.open(stream=annotated_pdf, filetype="pdf")
                        merged_doc  = fitz.open()
                        merged_doc.insert_pdf(parking_doc)
                        merged_doc.insert_pdf(toll_doc)
                        parking_doc.close(); toll_doc.close()
                        out_merged = io.BytesIO()
                        merged_doc.save(out_merged, garbage=4, deflate=True)
                        merged_bytes = out_merged.getvalue()
                        merged_size  = len(merged_bytes)
                        if merged_size > SIZE_LIMIT:
                            st.info(f"合併後 {merged_size/1024/1024:.1f}MB，開始降階壓縮...")
                            compressed = None
                            for quality in [85, 75, 60, 45]:
                                buf = io.BytesIO()
                                merged_doc.save(buf, garbage=4, deflate=True, deflate_images=True, deflate_fonts=True)
                                comp_doc = fitz.open(stream=buf.getvalue(), filetype="pdf")
                                out_comp = io.BytesIO()
                                writer_doc = fitz.open()
                                scale = 1.0
                                if quality <= 75: scale = 0.85
                                if quality <= 60: scale = 0.70
                                if quality <= 45: scale = 0.55
                                for pg in comp_doc:
                                    mat = fitz.Matrix(scale, scale)
                                    pix = pg.get_pixmap(width=pg.rect.width, height=pg.rect.height)
                                    img_pdf = fitz.open()
                                    img_page = img_pdf.new_page(width=pg.rect.width, height=pg.rect.height)
                                    img_page.insert_image(img_page.rect, pixmap=pix)
                                    writer_doc.insert_pdf(img_pdf)
                                writer_doc.save(out_comp, garbage=4, deflate=True)
                                result = out_comp.getvalue()
                                comp_doc.close(); writer_doc.close()
                                if len(result) <= SIZE_LIMIT:
                                    compressed = result
                                    break
                            st.session_state['merged_pdf'] = compressed or result
                            st.session_state['merged_compressed'] = True
                            st.session_state['merged_size'] = len(st.session_state['merged_pdf'])
                            st.session_state['merged_quality'] = quality
                        else:
                            st.session_state['merged_pdf'] = merged_bytes
                            st.session_state['merged_compressed'] = False
                            st.session_state['merged_size'] = merged_size
                        merged_doc.close()

                    st.success(f"✅ 完成！共比對 **{len(matched)}** 筆通行費")
                    unmatched = set(toll_map.keys()) - matched
                    if unmatched:
                        st.markdown(f'<div class="warn-box">⚠️ PDF有記錄但申請表未找到的日期：{", ".join(sorted(unmatched))}</div>', unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"錯誤：{e}")
                    import traceback; st.code(traceback.format_exc())

    if st.session_state.toll_excel or st.session_state.toll_pdf_out or st.session_state.mileage_pdf_out:
        dl1, dl2, dl3 = st.columns(3, gap="small")
        with dl1:
            if st.session_state.toll_excel:
                te_name = te_excel.name if te_excel else "T_E申請表.xlsx"
                st.download_button("💾 下載更新後的 Excel（含稽核頁籤）", st.session_state.toll_excel,
                    f"{selected_sheet}_對帳稽核_{te_name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with dl2:
            if st.session_state.toll_pdf_out and toll_pdf:
                st.download_button("💾 下載標註後的遠通電收", st.session_state.toll_pdf_out,
                    f"標註_{selected_sheet}_{toll_pdf.name}", mime="application/pdf")
        with dl3:
            if LIBREOFFICE_AVAILABLE:
                if st.session_state.mileage_pdf_out:
                    st.download_button("💾 下載原版格式里程 PDF", st.session_state.mileage_pdf_out,
                        f"{selected_sheet}_里程明細.pdf", mime="application/pdf")
            else:
                st.info("💡 雲端尚未啟動 LibreOffice，請下載 Excel 後在電腦另存 PDF。")

    if st.session_state.get('merged_pdf'):
        merged_data = st.session_state.get('merged_pdf') or b''
        size_mb = st.session_state.get('merged_size', len(merged_data)) / 1024 / 1024
        was_comp = st.session_state.get('merged_compressed', False)
        month_str = selected_sheet or datetime.now().strftime("%Y%m")
        label = f"✅ 壓縮完成：{size_mb:.1f}MB" if was_comp else f"✅ 合併完成：{size_mb:.1f}MB"
        st.markdown(f'<div class="success-box">{label}</div>', unsafe_allow_html=True)
        st.download_button(f"💾 下載合併PDF（{size_mb:.1f}MB）",
            data=st.session_state['merged_pdf'],
            file_name=f"{month_str}_停車費＋通行費.pdf",
            mime="application/pdf", type="primary")

    if st.session_state.audit_df is not None:
        with st.expander("🔍 檢視通行費對帳稽核報告"):
            st.markdown("**每日明細金額雙向稽核明細**")
            def highlight_diff(row):
                if row['狀態'] == '❌ 金額不符':
                    return ['background-color: #ffcccc'] * len(row)
                return ['background-color: #e6ffed'] * len(row)
            st.dataframe(st.session_state.audit_df.style.apply(highlight_diff, axis=1), use_container_width=True)
            c1, c2 = st.columns(2)
            c1.metric("Excel 總金額", f"{int(st.session_state.audit_df['Excel金額'].sum()):,} 元")
            c2.metric("遠通 PDF 總金額", f"{int(st.session_state.audit_df['PDF金額'].sum()):,} 元")

    if toll_pdf:
        with st.expander("🔍 預覽遠通電收原始解析結果"):
            toll_pdf.seek(0)
            pm = parse_toll_from_pdf(toll_pdf.read())
            if pm:
                st.markdown(f"**共 {len(pm)} 筆，合計 NT$ {sum(pm.values()):,} 元**")
                for d, a in sorted(pm.items()):
                    st.markdown(f"- {d}：{a} 元")
            else:
                st.warning("未解析到通行費資料")

    st.markdown("<div style='margin-top:2rem'></div>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">📱 電信費 PDF 處理</div>', unsafe_allow_html=True)

    telecom_file = st.file_uploader("上傳電信費 PDF", type="pdf", key="telecom_up")
    tc_col1, tc_col2 = st.columns([3, 2], gap="medium")
    with tc_col1:
        password = st.text_input("PDF 密碼（無密碼請留空）", type="password",
            placeholder="身分證末4碼 / 生日 MMDD", key="telecom_pwd")
        st.markdown("""<div class="warn-box" style="font-size:.82rem">
        💡 台灣大哥大／遠傳：身分證末4碼<br>
        中華電信：生日 MMDD 或身分證末4碼<br>
        亞太電信：出生年月日 YYYYMMDD
        </div>""", unsafe_allow_html=True)
    with tc_col2:
        st.markdown("<div style='margin-top:1.8rem'></div>", unsafe_allow_html=True)
        if telecom_file:
            if st.button("🔓 移除密碼並擷取第一頁", type="primary", key="run_telecom"):
                telecom_file.seek(0)
                raw = telecom_file.read()
                passwords_to_try = list(dict.fromkeys([password, "", "0000"]))
                success = False
                for pwd in passwords_to_try:
                    ok, result_bytes, msg = remove_pdf_password_and_extract_page1(raw, pwd)
                    if ok:
                        st.session_state.telecom_pdf = result_bytes
                        st.success(f"✅ {msg}")
                        if pwd != password:
                            st.info(f"使用密碼「{pwd}」成功解密")
                        success = True
                        break
                if not success:
                    st.error("❌ 密碼錯誤，請確認後重試")
        if st.session_state.telecom_pdf:
            orig = telecom_file.name.replace('.pdf', '') if telecom_file else "電信費"
            st.download_button("💾 下載（已解密，僅第一頁）",
                data=st.session_state.telecom_pdf,
                file_name=f"{orig}_第一頁.pdf", mime="application/pdf")
            st.markdown('<div class="success-box" style="font-size:.82rem">✅ 下載後直接上傳至 Concur 作為電信費附件</div>', unsafe_allow_html=True)


# ╔══════════════════════════════════════════╗
# ║  右側：加油費計算                        ║
# ╚══════════════════════════════════════════╝
with col_fuel:
    current_invoice_rows = []
    for i in range(1, 11):
        t_val = st.session_state.get(f"inv_t{i}", 0)
        x_val = st.session_state.get(f"inv_x{i}", 0)
        if t_val > 0:
            current_invoice_rows.append((t_val, x_val))

    if current_invoice_rows:
        temp_total_amount = sum(r[0] for r in current_invoice_rows)
        temp_total_tax = sum(r[1] for r in current_invoice_rows)
        m_allowance = st.session_state.get("mileage_manual", 0)
        temp_km = math.ceil(max(0, m_allowance - temp_total_amount) / 7) if m_allowance > 0 else 0
        st.session_state.fuel_amount = temp_total_amount
        st.session_state.fuel_tax = temp_total_tax
        st.session_state.mileage_distance = temp_km
    else:
        st.session_state.fuel_amount = 0
        st.session_state.fuel_tax = 0
        st.session_state.mileage_distance = 0

    st.markdown('<div class="section-title">📋 Concur 快速填寫對照表</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background:#F2F6FA;border-left:5px solid #1F4E79;padding:15px;border-radius:6px;margin-bottom:1.5rem;">
      <p style="margin:0 0 10px 0;font-weight:700;color:#1F4E79;font-size:0.95rem;">💡 複製上月申報單後，僅需更新以下 3 筆動態欄位：</p>
      <table style="width:100%;border-collapse:collapse;font-size:0.9rem;">
        <tr style="border-bottom:1px solid #e0e0e0;">
          <td style="padding:8px 0;font-weight:600;color:#333;">1. Personal Car Mileage</td>
          <td style="text-align:right;padding:8px 0;font-weight:700;color:#C00000;font-size:1.15rem;">
            {safe_format_num(st.session_state.mileage_distance)} <span style="font-size:0.8rem;font-weight:400;color:#555;">公里 (Distance)</span>
          </td>
        </tr>
        <tr style="border-bottom:1px solid #e0e0e0;">
          <td style="padding:8px 0;font-weight:600;color:#333;">2. Tolls/Road Charges/ Parking</td>
          <td style="text-align:right;padding:8px 0;font-weight:700;color:#1F4E79;font-size:1.15rem;">
            TWD {safe_format_num(st.session_state.tolls_parking_amount)} <span style="font-size:0.8rem;font-weight:400;color:#555;">(Amount)</span>
          </td>
        </tr>
        <tr>
          <td style="padding:8px 0;font-weight:600;color:#333;">3. Fuel</td>
          <td style="text-align:right;padding:8px 0;">
            <span style="font-weight:700;color:#333;font-size:1.15rem;">TWD {safe_format_num(st.session_state.fuel_amount)}</span> <span style="font-size:0.8rem;color:#555;">(Amount)</span><br>
            <span style="font-weight:700;color:#D35400;font-size:1.1rem;">TWD {safe_format_num(st.session_state.fuel_tax)}</span> <span style="font-size:0.8rem;color:#555;">(Tax Amount)</span>
          </td>
        </tr>
      </table>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title">⛽ 加油費計算</div>', unsafe_allow_html=True)

    if st.session_state.mileage_allowance:
        mileage_val = int(st.session_state.mileage_allowance)
        if st.session_state.get("mileage_manual", 0) != mileage_val:
            st.session_state["mileage_manual"] = mileage_val
        st.markdown(f'<div class="info-box">📊 里程津貼自動帶入：<b>NT$ {mileage_val:,}</b></div>', unsafe_allow_html=True)

    mileage_input = st.number_input("💰 總里程津貼（可手動修改）", min_value=0, step=100, key="mileage_manual")

    st.markdown("**🧾 加油發票**")
    fuel_pdf_file = st.file_uploader("上傳加油發票PDF（自動解析總計金額）", type="pdf", key="fuel_pdf_upload")

    if fuel_pdf_file:
        if st.button("🔍 自動解析發票金額", key="parse_fuel"):
            with st.spinner("解析中..."):
                fuel_pdf_file.seek(0)
                parsed = parse_fuel_pdf_totals(fuel_pdf_file.read())
            if parsed:
                for i, total in enumerate(parsed[:10], 1):
                    st.session_state[f"inv_t{i}"] = total
                    sales = round(total / 1.05)
                    st.session_state[f"inv_x{i}"] = round(sales * 0.05)
                for i in range(len(parsed[:10]) + 1, 11):
                    st.session_state[f"inv_t{i}"] = 0
                    st.session_state[f"inv_x{i}"] = 0
                over = "（超過10張，請分批上傳）" if len(parsed) > 10 else ""
                st.markdown(f'<div class="success-box">✅ 解析到 <b>{len(parsed)}</b> 筆發票：{parsed[:10]} {over}</div>', unsafe_allow_html=True)
                st.rerun()
            else:
                st.markdown('<div class="warn-box">⚠️ 未自動解析到金額，請手動輸入</div>', unsafe_allow_html=True)

    invoice_rows = []
    with st.expander("✍️ 手動微調 10 組發票金額（正常無須開啟）"):
        hc1, hc2 = st.columns([3, 2])
        with hc1: st.markdown("<div style='font-size:.8rem;color:#888;padding:2px 0'>發票總額</div>", unsafe_allow_html=True)
        with hc2: st.markdown("<div style='font-size:.8rem;color:#888;padding:2px 0'>稅額（可修改）</div>", unsafe_allow_html=True)
        for i in range(1, 11):
            ic1, ic2 = st.columns([3, 2])
            with ic1:
                total = st.number_input(f"總額{i}", min_value=0, step=1,
                    key=f"inv_t{i}", on_change=auto_tax, args=(i,), label_visibility="collapsed")
            with ic2:
                tax = st.number_input(f"稅額{i}", min_value=0, step=1,
                    key=f"inv_x{i}", label_visibility="collapsed")
            if total > 0:
                invoice_rows.append((total, tax))

    if invoice_rows:
        st.markdown("---")
        html_table, total_amount, total_tax, km, amt = build_results_html(invoice_rows, mileage_input)
        st.session_state.fuel_amount = total_amount
        st.session_state.fuel_tax = total_tax
        st.session_state.mileage_distance = km
        components.html(html_table, height=400, scrolling=False)
        st.markdown(f"""
        <div class="info-box" style="margin-top:.5rem">
        <b>Concur 填寫摘要</b><br>
        Fuel → Amount：<b>{total_amount:,}</b>　Tax Amount：<b>{total_tax:,}</b><br>
        Personal Car → Distance：<b>{km:,} 公里</b>（金額 {amt:,}）
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown('<div style="color:#aaa;text-align:center;padding:2rem 0;font-size:.9rem;">輸入發票金額後即時顯示結算表</div>', unsafe_allow_html=True)


save_persistent_state()
